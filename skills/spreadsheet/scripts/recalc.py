#!/usr/bin/env python3
"""Recalculate an .xlsx with LibreOffice and report what its formulas evaluate to.

    python3 scripts/recalc.py workbook.xlsx [--timeout 60] [--strict]

WHY THIS EXISTS
---------------
openpyxl writes formulas as strings and stores no cached result for them. Until
something recalculates the file, every formula cell reads back as empty:

    ws["B1"] = "=SUM(A1:A3)"; wb.save(path)
    load_workbook(path, data_only=True)["B1"].value   ->  None
    pd.read_excel(path)                               ->  the column is absent

So a workbook that looks correct in your code can arrive at the user with blank
columns. This runs LibreOffice over it (which computes every formula and writes
the results in), rewrites the file in place, then inspects the outcome.

WHAT IT REPORTS
---------------
JSON on stdout:

    status              ok | errors_found | unevaluated | failed
    total_formulas      formula cells found
    errors              cells whose value is an Excel error (#REF!, #NAME!, ...)
    unevaluated         formula cells still empty AFTER recalculation — usually a
                        function this LibreOffice build cannot compute
    by_type             {"#NAME?": ["Sheet1!B4", ...], ...}, capped per type

EXIT CODES
----------
0  unless the recalculation itself could not run (status "failed"), or --strict
   was passed and the workbook has errors or unevaluated formulas.

Without --strict a clean exit does NOT mean a clean workbook — read `status`.
And note that even status "ok" only proves the formulas *evaluate*; a range that
is off by one evaluates perfectly and is still wrong.
"""

# Keeps `str | None` and `dict[str, list[str]]` annotations working on 3.7–3.9,
# so the script isn't tied to whatever Python the sandbox image ships.
from __future__ import annotations

import argparse
import glob
import json
import os
import shutil
import subprocess
import sys
import tempfile

# Excel's error sentinels, as LibreOffice writes them back into the file.
ERROR_VALUES = {
    "#REF!",
    "#NAME?",
    "#VALUE!",
    "#DIV/0!",
    "#N/A",
    "#NULL!",
    "#NUM!",
    "#SPILL!",
    "#CALC!",
    "#GETTING_DATA",
}

MAX_LOCATIONS_PER_TYPE = 100


def find_soffice() -> str | None:
    for name in ("soffice", "libreoffice"):
        path = shutil.which(name)
        if path:
            return path
    # macOS bundle, for local development.
    bundled = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    return bundled if os.path.isfile(bundled) else None


def recalculate(path: str, timeout: int) -> tuple[bool, str]:
    """Convert xlsx -> xlsx via LibreOffice, which recalculates on load.

    Returns (ok, message). The file is replaced in place on success.
    """
    soffice = find_soffice()
    if not soffice:
        return False, (
            "LibreOffice not found. This skill needs the nc-chat sandbox image; "
            "do not attempt to install it."
        )

    outdir = tempfile.mkdtemp(prefix="recalc-")
    # A private profile dir: a sandbox often has no writable HOME, and without
    # this soffice exits silently having written nothing.
    profile = tempfile.mkdtemp(prefix="lo-profile-")

    try:
        result = subprocess.run(
            [
                soffice,
                f"-env:UserInstallation=file://{profile}",
                "--headless",
                "--norestore",
                "--convert-to",
                "xlsx",
                "--outdir",
                outdir,
                os.path.abspath(path),
            ],
            capture_output=True,
            text=True,
            timeout=timeout,
        )
    except subprocess.TimeoutExpired:
        shutil.rmtree(outdir, ignore_errors=True)
        shutil.rmtree(profile, ignore_errors=True)
        return False, f"LibreOffice timed out after {timeout}s"

    produced = glob.glob(os.path.join(outdir, "*.xlsx"))
    if result.returncode or not produced:
        shutil.rmtree(outdir, ignore_errors=True)
        shutil.rmtree(profile, ignore_errors=True)
        detail = (result.stderr or result.stdout or "").strip()[-400:]
        return False, f"LibreOffice failed: {detail or 'no output produced'}"

    shutil.copyfile(produced[0], path)
    shutil.rmtree(outdir, ignore_errors=True)
    shutil.rmtree(profile, ignore_errors=True)
    return True, "recalculated"


def inspect(path: str) -> dict:
    """Two passes: formulas from one load, computed values from the other."""
    try:
        import openpyxl
    except ImportError:
        return {"status": "failed", "error": "openpyxl is not available"}

    formulas_wb = openpyxl.load_workbook(path)
    values_wb = openpyxl.load_workbook(path, data_only=True)

    total_formulas = 0
    errors = 0
    unevaluated = 0
    by_type: dict[str, list[str]] = {}
    truncated: dict[str, int] = {}

    def note(kind: str, location: str) -> None:
        bucket = by_type.setdefault(kind, [])
        if len(bucket) < MAX_LOCATIONS_PER_TYPE:
            bucket.append(location)
        else:
            truncated[kind] = truncated.get(kind, 0) + 1

    for name in formulas_wb.sheetnames:
        fsheet = formulas_wb[name]
        vsheet = values_wb[name]

        for row in fsheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value.startswith("="):
                    continue

                total_formulas += 1
                location = f"{name}!{cell.coordinate}"
                computed = vsheet[cell.coordinate].value

                if isinstance(computed, str) and computed.strip() in ERROR_VALUES:
                    errors += 1
                    note(computed.strip(), location)
                elif computed is None:
                    # Recalculation ran but produced nothing for this cell.
                    # A formula returning "" also lands here — check the ones
                    # named before assuming they are all broken.
                    unevaluated += 1
                    note("unevaluated", location)

    if errors:
        status = "errors_found"
    elif unevaluated:
        status = "unevaluated"
    else:
        status = "ok"

    report = {
        "status": status,
        "total_formulas": total_formulas,
        "errors": errors,
        "unevaluated": unevaluated,
        "by_type": by_type,
    }
    if truncated:
        report["locations_truncated"] = truncated
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("path", help="Path to the .xlsx to recalculate in place.")
    parser.add_argument("--timeout", type=int, default=60)
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Exit non-zero when the workbook has errors or unevaluated formulas.",
    )
    args = parser.parse_args()

    if not os.path.isfile(args.path):
        print(json.dumps({"status": "failed", "error": f"no such file: {args.path}"}))
        return 1

    ok, message = recalculate(args.path, args.timeout)
    if not ok:
        print(json.dumps({"status": "failed", "error": message}, indent=2))
        return 1

    report = inspect(args.path)
    print(json.dumps(report, indent=2))

    if report["status"] == "failed":
        return 1
    if args.strict and report["status"] != "ok":
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
